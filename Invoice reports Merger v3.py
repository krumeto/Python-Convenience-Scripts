#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np
import time
import openpyxl as xl
get_ipython().run_line_magic('matplotlib', 'inline')


# In[2]:


def customer_ibr_reader(link):
    """
    A function to read customer IBRs to a clean pandas dataframe

    """
    ibr_customer = pd.read_excel(link, skiprows=10, encoding = "ISO-8859-1")
    if ["Serial Number" in ibr_customer.columns.unique()]==False:
        ibr_customer = bad_ibr(link)
    ibr_customer = bad_ibr(link)
    ibr_customer.columns = [str(x).lower() for x in ibr_customer.columns]
    ibr_customer = ibr_customer[:-2]
    header = pd.read_excel(link, encoding = "ISO-8859-1", nrows=3, header = None, index = False)
    header.set_index(0, inplace=True)
    ibr_customer['customer name'] = header.T.iloc[0][0]
    ibr_customer['amp id'] = header.T.iloc[0][2]
    columns = ['Customer Name', 'Amp Id', "Product Number", 'Product Description','Serial Number', "Base charge", "Count Mono", "Mono Rate", "Total Mono", "Count Color", "Color Rate", "Total Color", "Total Charge per device", "Tax Amount", "Total Amount", "Count Color Professional", "Color Professional Rate", "Total Color Professional", "Count Color Accent", "Color Accent Rate", "Total Color Accent"]
    columns = [x.lower() for x in columns]
    for col in columns:
        if col not in ibr_customer.columns.unique():
            ibr_customer[col] = np.nan
    
    cleaned_ibr = ibr_customer[columns]
    return cleaned_ibr


# In[3]:


def bad_ibr(link):
    ibr_customer = pd.read_excel(link, encoding = "ISO-8859-1")
    index = list(np.where(ibr_customer.iloc[:,0]=="Product Number")[0])
    if len(index) == 0:
        index = list(np.where(ibr_customer.iloc[:,0]=="Customer Name")[0])
    return pd.read_excel(link, skiprows=index[0]+1, encoding = "ISO-8859-1") 


# In[4]:


def customer_ibr_merger():
    tic = time.time()
    customer_columns = ['Customer Name', 'Amp Id', "Product Number", 'Product Description','Serial Number', "Base charge", "Count Mono", "Mono Rate", "Total Mono", "Count Color", "Color Rate", "Total Color", "Total Charge per device", "Tax Amount", "Total Amount", "Count Color Professional", "Color Professional Rate", "Total Color Professional", "Count Color Accent", "Color Accent Rate", "Total Color Accent"]
    customer_columns = [x.lower() for x in customer_columns]
    
    ibr = pd.DataFrame(columns=customer_columns)
    import os
    file_list = [f for f in os.listdir('.') if os.path.isfile(os.path.join('.', f))]
    excel_files = []
    for file in file_list:
        filename = os.fsdecode(file)
        if filename.endswith(".xlsx"):
            excel_files.append(filename)
    
    for excel in excel_files:
        print(excel)
        var_ibr = customer_ibr_reader(excel)
        ibr = pd.concat([ibr, var_ibr])
    
    tac = time.time()
    print(tac - tic)
    return ibr.to_excel("Combined IBR.xlsx")


# In[5]:


customer_ibr_merger()


# In[4]:


pd.read_excel("9TB2040_May_2019.xlsx")


# In[21]:


import pandas as pd
header = pd.read_csv('Copy of 0040_April_2019.xlsx', skiprows=10, encoding = "ISO-8859-1")
#test_df = pd.read_excel('Copy of 0040_April_2019.xlsx', skiprows=10)
#header.set_index(0, inplace=True)


# In[39]:


import openpyxl
theFile = openpyxl.load_workbook('0040_April_2019.xlsx', keep_vba=True)
allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))


# In[26]:


def find_specific_cell():
    for row in range(1, currentSheet.max_row + 1):
        for column in "ABCD":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if currentSheet[cell_name].value == "Customer Name":
                #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                return cell_name


# In[27]:


for sheet in allSheetNames:
    print("Current sheet name is {}" .format(sheet))
    currentSheet = theFile[sheet]
    print(find_specific_cell())


# In[28]:


ws = theFile.active


# In[36]:


for row in ws.iter_rows(range_string="B2:B8", max_row=2):
    for cell in row:
        print(cell.value)

